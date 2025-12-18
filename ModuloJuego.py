# ModuloJuego.py

import random
from datetime import datetime
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
import matplotlib.pyplot as plt

EXCEL_PATH = "./Estadisticas.xlsx"

# Funciones de Excel

def crear_excel():
    """Crea el archivo Excel si no existe, con cabeceras."""
    try:
        load_workbook(EXCEL_PATH)
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
        hoja = workbook.active
        hoja.title = "Estadísticas"
        cabeceras = ["Fecha", "Jugador", "Resultado", "Dificultad", "Modo", "Intentos"]
        for col, cabecera in enumerate(cabeceras, start=1):
            cell = hoja.cell(row=1, column=col, value=cabecera)
            cell.font = Font(bold=True)
            hoja.column_dimensions[get_column_letter(col)].width = 20
        workbook.save(EXCEL_PATH)

def guardar_resultado(nombre, resultado, dificultad, modo, intentos_usados):
    """Guarda los resultados de una partida en Excel."""
    crear_excel()
    workbook = load_workbook(EXCEL_PATH)
    hoja = workbook["Estadísticas"]
    fecha_actual = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    # Siempre rellenar las 6 columnas
    fila = [
        fecha_actual,
        nombre if nombre else "Desconocido",
        resultado if resultado else "Desconocido",
        dificultad if dificultad else 1,
        modo if modo else "Solitario",
        intentos_usados if intentos_usados else 0
    ]
    hoja.append(fila)
    workbook.save(EXCEL_PATH)

# Funciones de juego

def elegir_dificultad():
    while True:
        try:
            opcion = int(input("Elige dificultad (1-Fácil, 2-Medio, 3-Difícil): "))
            if 1 <= opcion <= 3:
                return opcion
            print("Número fuera de rango")
        except ValueError:
            print("Introduce un número válido")

def generar_numero(dificultad, modo):
    if dificultad == 1:
        maximo = 1000
        intentos = 20
    elif dificultad == 2:
        maximo = 1000
        intentos = 12
    else:
        maximo = 1000
        intentos = 5

    if modo == "solitario":
        numero = random.randint(1, maximo)
    else:  # 2 jugadores
        while True:
            try:
                numero = int(input(f"Jugador 1, introduce el número a adivinar (1-{maximo}): "))
                if 1 <= numero <= maximo:
                    break
                print("Número fuera de rango")
            except ValueError:
                print("Número no válido")
    return numero, maximo, intentos

def adivinar(numero_secreto, maximo, intentos):
    no_acertado = []
    for i in range(1, intentos+1):
        while True:
            try:
                guess = int(input(f"Intento {i}/{intentos}, introduce un número entre 1 y {maximo}: "))
                if 1 <= guess <= maximo:
                    break
                print("Número fuera de rango")
            except ValueError:
                print("Número no válido")
        if guess == numero_secreto:
            print("¡Has acertado!")
            return True, i
        elif guess < numero_secreto:
            print("El número es mayor")
        else:
            print("El número es menor")
        no_acertado.append(guess)
    print(f"Has perdido. El número era {numero_secreto}")
    return False, intentos

# Funciones de gráficos

def grafico_resultados():
    crear_excel()
    workbook = load_workbook(EXCEL_PATH)
    hoja = workbook["Estadísticas"]
    resultados = {"Ganado":0, "Perdido":0}
    for fila in hoja.iter_rows(min_row=2, values_only=True):
        if fila[2] in resultados:  # Ignorar filas vacías
            resultados[fila[2]] += 1
    plt.bar(resultados.keys(), resultados.values(), color=["green","red"])
    plt.title("Resultados del juego")
    plt.show()

def grafico_dificultades():
    crear_excel()
    workbook = load_workbook(EXCEL_PATH)
    hoja = workbook["Estadísticas"]
    dificultades = {"Fácil":0, "Medio":0, "Difícil":0}
    for fila in hoja.iter_rows(min_row=2, values_only=True):
        dif = fila[3]
        if dif == 1:
            dificultades["Fácil"] += 1
        elif dif == 2:
            dificultades["Medio"] += 1
        elif dif == 3:
            dificultades["Difícil"] += 1
    plt.bar(dificultades.keys(), dificultades.values(), color=["green","orange","red"])
    plt.title("Dificultad de partidas jugadas")
    plt.show()
